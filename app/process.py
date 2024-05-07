import re
import openpyxl
from app.models import Entry
from django.http import HttpResponse


def sheet_activation(workbook, sheet):
    workbook = openpyxl.load_workbook(workbook)
    sheet = workbook[sheet]
    workbook.active = sheet
    sheet = workbook.active

    return sheet

def gather_clubbed_name(master_workbook):
    insurer_data = {}
    master_sheet = sheet_activation(master_workbook, "name")
    
    for row in range(2, master_sheet.max_row+1):
        company_insurer = master_sheet.cell(row=row, column=1)
        company_name = master_sheet.cell(row=row, column=2)
        company_clubbed_name = master_sheet.cell(row=row, column=3)
        insurer_data[company_insurer.value.strip()] = [company_name.value, company_clubbed_name.value]

    return insurer_data

def lob_products(master_workbook):
    lob_products = []
    master_sheet = sheet_activation(master_workbook, "lob")

    for row in range(1, master_sheet.max_row+1):
        product_name = master_sheet.cell(row=row, column=1).value.strip()
        lob_products.append(product_name)

    return lob_products

def insurer_category_collection(master_workbook):
    company_categories = {}
    master_sheet = sheet_activation(master_workbook, "category")

    for row in range(2, master_sheet.max_row+1):
        company_clubbed_name = master_sheet.cell(row=row, column=1).value.strip()
        category = master_sheet.cell(row=row, column=2).value.strip()
        company_categories[company_clubbed_name] = category

    return company_categories

def extract_data_from_sheets(task_sheet, master_workbook, year, month):
    try:
        product_list = []
        for col in range(2, task_sheet.max_column + 1):
            cell_value = task_sheet.cell(row=2, column=col).value
            if cell_value is not None:
                product_list.append(cell_value.strip())

        company_list = [task_sheet.cell(row=row, column=1).value.strip() for row in range(4, task_sheet.max_row+1) if task_sheet.cell(row=row, column=1).value is not None]

        company_name_data = gather_clubbed_name(master_workbook=master_workbook)
        lob_product_list = lob_products(master_workbook=master_workbook)
        company_category_data = insurer_category_collection(master_workbook=master_workbook)

        company_name_list = list(company_name_data.keys())
        company_data = []
        for row in range(4, task_sheet.max_row+1):
            company_name = task_sheet.cell(row=row, column=1).value
            if company_name and company_name in company_name_list:
                for index, product in enumerate(product_list):
                    if product in lob_product_list:
                        company_value = task_sheet.cell(row=row, column=index+2).value
                        company_year = year
                        company_month = month
                        company_clubbed_name = company_name_data[company_name][1]
                        company_category = company_category_data[company_clubbed_name]
                        company_object = Entry(
                            year=company_year,
                            month=company_month[:3],
                            category=company_category,
                            clubbed_name=company_clubbed_name,
                            product=product,
                            value=company_value
                        )
                        company_object.save()

        return company_data
    except Exception as e:
        return None

def process_input_data(task_workbook, master_workbook):
    try:
        sheet_names = ['Miscellaneous portfolio', 'Segmentwise Report']
        for sheet_name in sheet_names:
            task_sheet = sheet_activation(task_workbook, sheet_name)
            sheet_title = task_sheet.cell(row=1, column=1).value.strip()
            match = re.search(r'(January|February|March|April|May|June|July|August|September|October|November|December)\s+(\d{4})', sheet_title)
            if match:
                month = match.group(1)
                year = match.group(2)
            else:
                month, year = 0, 0
            result = extract_data_from_sheets(task_sheet=task_sheet, master_workbook=master_workbook, year=year, month=month)
        
        if result:
            return result

    except Exception as e:
        return e

def generate_excel_output():
    output_headers = ["Year", "Month", "category", "clubbed_name", "Product", "Value"]
    output_excel = openpyxl.Workbook()
    output_sheet = output_excel.active

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="system-output.xlsx"'

    for index, title in enumerate(output_headers):
        output_sheet.cell(row=1, column=index+1).value = title

    all_entries = list(Entry.objects.all().order_by('clubbed_name', 'year', 'product').values(
        'year',
        'month',
        'category',
        'clubbed_name',
        'product',
        'value'
    ))

    for index, entry in enumerate(all_entries):
        output_sheet.cell(row=index+2, column=1).value = entry['year']
        output_sheet.cell(row=index+2, column=2).value = entry['month']
        output_sheet.cell(row=index+2, column=3).value = entry['category']
        output_sheet.cell(row=index+2, column=4).value = entry['clubbed_name']
        output_sheet.cell(row=index+2, column=5).value = entry['product']
        output_sheet.cell(row=index+2, column=6).value = entry['value']

    output_excel.save(response)
    return response